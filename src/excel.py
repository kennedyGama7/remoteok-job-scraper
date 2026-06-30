from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def salvar_excel(lista_vagas, tecnologia):
    planilha = Workbook()
    aba = planilha.active
    aba.title = "Vagas"

    aba.append([
        "Título",
        "Empresa",
        "Localização",
        "Link",
        "Tags"
    ])

    for celula in aba[1]:
        celula.font = Font(
            bold=True,
            color="FFFFFF"
        )
        celula.fill = PatternFill(
            fill_type="solid",
            start_color="4F81BD",
            end_color="4F81BD"
       )
    aba.freeze_panes = "A2"

    for vaga in lista_vagas:
        aba.append([
            vaga.titulo,
            vaga.empresa,
            vaga.localizacao,
            vaga.link,
            vaga.tags
    ])
       

        linha = aba.max_row

        celula_link = aba.cell(row=linha, column=4)

        celula_link.hyperlink = vaga.link
        celula_link.style = "Hyperlink" 

    for coluna in aba.columns:
        maior_tamanho = 0
        letra_coluna = get_column_letter(coluna[0].column)

        for celula in coluna:
            if celula.value is not None:
                tamanho = len(str(celula.value))
                if tamanho > maior_tamanho:
                    maior_tamanho = tamanho

        aba.column_dimensions[letra_coluna].width = maior_tamanho + 2

    ultima_linha = aba.max_row
    ultima_coluna = get_column_letter(aba.max_column)
    intervalo = f"A1:{ultima_coluna}{ultima_linha}"

    tabela = Table(
        displayName="TabelaVagas",
        ref=intervalo
            )
    
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
        )
    
    tabela.tableStyleInfo = estilo

    aba.add_table(tabela)

    planilha.save(f"output/vagas_{tecnologia.lower()}.xlsx")